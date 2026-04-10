import os
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from typing import List, Dict
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Processes various document formats and extracts text for embedding"""
    
    def __init__(self, knowledge_base_path: str):
        self.knowledge_base_path = Path(knowledge_base_path)
    
    def extract_text_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF file"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF {file_path}: {e}")
            return ""
    
    def extract_text_from_docx(self, file_path: Path) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {file_path}: {e}")
            return ""
    
    def extract_text_from_xlsx(self, file_path: Path) -> str:
        """Extract text from XLSX file"""
        try:
            workbook = load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    text += row_text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from XLSX {file_path}: {e}")
            return ""
    
    def process_document(self, file_path: Path) -> Dict[str, str]:
        """Process a document and return its content"""
        file_ext = file_path.suffix.lower()
        file_name = file_path.name
        
        text = ""
        if file_ext == '.pdf':
            text = self.extract_text_from_pdf(file_path)
        elif file_ext == '.docx' or file_ext == '.doc':
            text = self.extract_text_from_docx(file_path)
        elif file_ext == '.xlsx' or file_ext == '.xls':
            text = self.extract_text_from_xlsx(file_path)
        else:
            logger.warning(f"Unsupported file type: {file_ext}")
            return None
        
        if not text.strip():
            logger.warning(f"No text extracted from {file_name}")
            return None
        
        return {
            "filename": file_name,
            "content": text,
            "file_path": str(file_path)
        }
    
    def process_all_documents(self) -> List[Dict[str, str]]:
        """Process all documents in the knowledge base directory"""
        documents = []
        
        if not self.knowledge_base_path.exists():
            logger.error(f"Knowledge base path does not exist: {self.knowledge_base_path}")
            return documents
        
        for file_path in self.knowledge_base_path.iterdir():
            if file_path.is_file():
                logger.info(f"Processing document: {file_path.name}")
                doc_data = self.process_document(file_path)
                if doc_data:
                    documents.append(doc_data)
        
        logger.info(f"Processed {len(documents)} documents")
        return documents

